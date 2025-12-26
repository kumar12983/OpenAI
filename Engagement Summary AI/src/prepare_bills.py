# -*- coding: utf-8 -*-
"""
Prepare Bills*.xlsx Input File
-------------------------------
Automates data preparation for Bills file before running engagement analysis.

Steps:
1. Add "Billing Amount" = "Total Invoice Amount incl Tax" - "Tax" in Export sheet
2. Add "Invoice Month" = TEXT(Invoice Date, "YYYY-MM") next to Invoice Date
3. Extract "Engagement ID" from "Lead Engagement Name (ID) Currency" column
4. Create pivot table "Billing" with Engagement ID, Billing Partner, and billing totals

Usage:
    python prepare_bills.py --input Bills_08.12.2025.xlsx --output Bills_08.12.2025_prepared.xlsx
"""

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def extract_engagement_id(name_str: str) -> str:
    """Extract Engagement ID from 'Lead Engagement Name (ID) Currency' format.
    
    Example: "Project ABC (E-12345678) USD" -> "E-12345678"
    """
    if pd.isna(name_str):
        return None
    
    # Look for pattern (E-XXXXXXXX) or similar
    match = re.search(r'\(([A-Z]-\d+)\)', str(name_str))
    if match:
        return match.group(1)
    return None


def prepare_bills_export(input_file: str, output_file: str, invoice_month_from: str = None) -> str:
    """Prepare Bills Export sheet with required columns and calculations."""
    
    print(f"Loading {input_file}...")
    
    # Read Export sheet
    try:
        df = pd.read_excel(input_file, sheet_name='Export', engine='openpyxl')
    except Exception as e:
        print(f"Error reading Export sheet: {e}", file=sys.stderr)
        sys.exit(1)
    
    # Strip column names
    df.columns = [str(c).strip() for c in df.columns]
    
    print(f"Found {len(df)} rows in Export sheet")
    
    # 1. Add Billing Amount column
    if 'Total Invoice Amount incl Tax' not in df.columns or 'Tax' not in df.columns:
        print("Warning: 'Total Invoice Amount incl Tax' or 'Tax' column not found. Skipping Billing Amount calculation.")
    else:
        df['Billing Amount'] = pd.to_numeric(df['Total Invoice Amount incl Tax'], errors='coerce').fillna(0) - \
                                pd.to_numeric(df['Tax'], errors='coerce').fillna(0)
        print("✓ Added 'Billing Amount' column")
    
    # 2. Add Invoice Month column
    if 'Invoice Date' not in df.columns:
        print("Warning: 'Invoice Date' column not found. Skipping Invoice Month calculation.")
    else:
        df['Invoice Date'] = pd.to_datetime(df['Invoice Date'], errors='coerce')
        df['Invoice Month'] = df['Invoice Date'].dt.strftime('%Y-%m')
        print("✓ Added 'Invoice Month' column")
    
    # 3. Extract Engagement ID
    if 'Lead Engagement Name (ID) Currency' not in df.columns:
        print("Warning: 'Lead Engagement Name (ID) Currency' column not found. Skipping Engagement ID extraction.")
    else:
        df['Engagement ID'] = df['Lead Engagement Name (ID) Currency'].apply(extract_engagement_id)
        extracted_count = df['Engagement ID'].notna().sum()
        print(f"✓ Extracted Engagement ID for {extracted_count}/{len(df)} rows")
    
    # 4. Create Billing pivot table
    if 'Engagement ID' in df.columns and 'Billing Partner' in df.columns:
        # Filter rows with valid Engagement ID
        df_valid = df[df['Engagement ID'].notna()].copy()
        
        # Filter by Invoice Month if specified
        if invoice_month_from and 'Invoice Month' in df_valid.columns:
            df_valid = df_valid[df_valid['Invoice Month'] >= invoice_month_from].copy()
            print(f"✓ Filtered to invoices from {invoice_month_from} onwards: {len(df_valid)} rows")
        
        # Group by Engagement ID and Billing Partner
        pivot_cols = ['Engagement ID', 'Billing Partner']
        agg_dict = {}
        
        if 'Total Invoice Amount incl Tax' in df.columns:
            agg_dict['Total Invoice Amount incl Tax'] = 'sum'
        if 'Tax' in df.columns:
            agg_dict['Tax'] = 'sum'
        if 'Billing Amount' in df.columns:
            agg_dict['Billing Amount'] = 'sum'
        
        if agg_dict:
            billing_pivot = df_valid.groupby(pivot_cols, as_index=False).agg(agg_dict)
            billing_pivot = billing_pivot.round(2)
            print(f"✓ Created Billing pivot with {len(billing_pivot)} engagement-partner combinations")
            
            # Calculate total billing
            if 'Billing Amount' in billing_pivot.columns:
                total_billing = billing_pivot['Billing Amount'].sum()
                print(f"  Total Billing Amount: {total_billing:,.2f}")
                print(f"  → Use --billings {total_billing/1_000_000:.2f}M in your analysis command")
        else:
            billing_pivot = None
            print("Warning: Could not create pivot - missing required columns")
    else:
        billing_pivot = None
        print("Warning: Missing Engagement ID or Billing Partner - cannot create pivot")
    
    # Write to output file
    print(f"\nWriting to {output_file}...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write updated Export sheet
        df.to_excel(writer, sheet_name='Export', index=False)
        
        # Write Billing pivot if available
        if billing_pivot is not None:
            billing_pivot.to_excel(writer, sheet_name='Billing', index=False)
        
        # Copy other sheets from original file
        try:
            original_wb = load_workbook(input_file, read_only=False)
            for sheet_name in original_wb.sheetnames:
                if sheet_name != 'Export':
                    original_sheet = original_wb[sheet_name]
                    new_sheet = writer.book.create_sheet(sheet_name)
                    
                    for row in original_sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet[cell.coordinate]
                            new_cell.value = cell.value
                            if cell.has_style:
                                new_cell.font = cell.font.copy()
                                new_cell.border = cell.border.copy()
                                new_cell.fill = cell.fill.copy()
                                new_cell.number_format = cell.number_format
                                new_cell.protection = cell.protection.copy()
                                new_cell.alignment = cell.alignment.copy()
        except Exception as e:
            print(f"Warning: Could not copy other sheets: {e}")
    
    print(f"✓ Successfully prepared {output_file}")
    
    # Return suggested billings parameter
    if billing_pivot is not None and 'Billing Amount' in billing_pivot.columns:
        total = billing_pivot['Billing Amount'].sum()
        return f"{total/1_000_000:.2f}M"
    return None


def main():
    parser = argparse.ArgumentParser(description='Prepare Bills*.xlsx input file for engagement analysis')
    parser.add_argument('--input', required=True, help='Input Bills Excel file (e.g., Bills_08.12.2025.xlsx)')
    parser.add_argument('--output', required=True, help='Output prepared Excel file (e.g., Bills_08.12.2025_prepared.xlsx)')
    parser.add_argument('--invoice-month-from', default='2025-08', help='Filter invoices from this month onwards (format: YYYY-MM, default: 2025-08)')
    
    args = parser.parse_args()
    
    if not Path(args.input).exists():
        print(f"Error: Input file '{args.input}' not found", file=sys.stderr)
        sys.exit(1)
    
    billings_param = prepare_bills_export(args.input, args.output, args.invoice_month_from)
    
    print("\n" + "="*60)
    print("NEXT STEPS:")
    print("="*60)
    if billings_param:
        print(f"1. Use this file in your analysis: --bills {args.output}")
        print(f"2. Use this billings parameter: --billings {billings_param}")
    else:
        print(f"1. Use this file in your analysis: --bills {args.output}")
        print("2. Check the 'Billing' sheet for total Billing Amount to use as --billings parameter")


if __name__ == '__main__':
    main()
