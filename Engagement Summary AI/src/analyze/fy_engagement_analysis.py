# -*- coding: utf-8 -*-
"""
FY Engagement & Employee Summary + KPI Bridge
------------------------------------------------
Automates end-to-end engagement analysis for a fiscal period using a WIP Excel extract.

Business rules implemented:
1) Use **Accounting Date** to filter rows to [FY_START..FY_END]. If missing, use **Transaction Date**.
2) Treat **Expense Amount** as **pass-through revenue**.
3) For each row, then aggregate by **Engagement ID + Engagement Name**:
   - ANSR = "ANSR / Tech Revenue"
   - Margin Cost = "Margin Cost"
   - TER = ANSR + Expense Amount
   - Margin Amount = ANSR - Margin Cost
   - Margin % (on ANSR) = (Margin Amount ÷ ANSR) × 100

Employee summary:
- Aggregates by employee identity, includes Hours, NSR, ANSR, Margin Cost, Expense Amount, #Engagements, #Opportunities, TER,
  Margin Amount, Margin % (on ANSR), EAF (ANSR/NSR).
- Adds derived "Level" from Rank/Grade mapping:
   Rank: Partner/Principal = 7
   Executive Director = 6
   Senior Manager & Grade 2 = 5
   Senior Manager & Grade 1 = 4
   Manager = 3
   Senior = 2
   Staff/Assistant = 1

KPI Bridge:
1) Parse BILLINGS (supports "20.11M" or plain numbers). Compute **Write-on** = BILLINGS − (sum of TER).
2) **Revised Margin Amount** = (sum of Margin Amount) + Write-on.
3) **Revised Revenue** = (sum of ANSR) + Write-on.
4) **Revised Margin %** = Revised Margin Amount ÷ Revised Revenue × 100.
5) **Additional Revenue Needed** to reach TARGET_MARGIN_PCT with same cost:
   Solve x from (MarginAmount_total + x) / (ANSR_total + x) = TARGET_MARGIN_PCT%.

Outputs:
- Excel workbook with sheets: "Engagement Summary", "Employee Summary", "Monthly Summary", "KPI Totals", "KPI Bridge".
- Optional printed Markdown table (Engagement Summary by TER desc).

Usage (CLI):
    python fy_engagement_analysis.py \
        --input WIPs_20251225.xlsx \
        --sheet Detail \
        --fy-start 2025-07-01 \
        --fy-end 2026-06-30 \
        --billings 17.9M \
        --target-margin-pct 29 \
        --bills Bills_08.12.2025.xlsx \
        --bob BoB_08.12.2025.xlsx \
        --output Engagement_Summary_FY26.xlsx \
        --print-markdown

Requires: pandas (with openpyxl engine), numpy.
"""

import argparse
import sys
from typing import Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def parse_date_safe(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors='coerce')


def parse_billings(s: str) -> float:
    t = s.strip().upper().replace(',', '')
    if t.endswith('M'):
        return float(t[:-1]) * 1_000_000
    return float(t)


def detect_header_row(df_raw: pd.DataFrame) -> Optional[int]:
    for i in range(len(df_raw)):
        row_vals = df_raw.iloc[i].astype(str).tolist()
        if 'Opportunity ID' in row_vals and 'Engagement ID' in row_vals:
            return i
    return None


def load_detail_frame(input_file: str, sheet_name: str, header_row_index: Optional[int]) -> pd.DataFrame:
    raw = pd.read_excel(input_file, sheet_name=sheet_name, header=None, engine='openpyxl')
    if header_row_index is None:
        header_row_index = detect_header_row(raw)
        if header_row_index is None:
            raise ValueError('Header row not found. Provide --header-row-index or ensure the sheet contains the expected headers.')
    headers = raw.iloc[header_row_index].tolist()
    df = raw.iloc[header_row_index + 1:].copy()
    df.columns = headers
    return df


def filter_fiscal_year(df: pd.DataFrame, fy_start: pd.Timestamp, fy_end: pd.Timestamp) -> pd.DataFrame:
    for c in ['Accounting Date', 'Transaction Date']:
        if c not in df.columns:
            df[c] = pd.NaT
    df['Accounting Date'] = parse_date_safe(df['Accounting Date'])
    df['Transaction Date'] = parse_date_safe(df['Transaction Date'])
    mask = (
        (df['Accounting Date'].notna() & (df['Accounting Date'] >= fy_start) & (df['Accounting Date'] <= fy_end)) |
        (df['Accounting Date'].isna() & df['Transaction Date'].notna() & (df['Transaction Date'] >= fy_start) & (df['Transaction Date'] <= fy_end))
    )
    return df.loc[mask].copy()


def coerce_numeric(df: pd.DataFrame, cols) -> pd.DataFrame:
    for c in cols:
        if c not in df.columns:
            df[c] = 0
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    return df


def engagement_summary(df: pd.DataFrame,
                      partners: Optional[pd.DataFrame] = None,
                      nui_etd: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    grp = df.groupby(['Engagement ID', 'Engagement Name']).agg({
        'ANSR / Tech Revenue': 'sum',
        'Margin Cost': 'sum',
        'Expense Amount': 'sum',
        'Charged Hours / Quantity': 'sum'
    }).reset_index()
    grp.rename(columns={'Charged Hours / Quantity': 'Hours'}, inplace=True)
    
    grp['TER'] = grp['ANSR / Tech Revenue'] + grp['Expense Amount']
    grp['Margin Amount'] = grp['ANSR / Tech Revenue'] - grp['Margin Cost']
    grp['Margin %'] = np.where(grp['ANSR / Tech Revenue'] != 0,
                               (grp['Margin Amount'] / grp['ANSR / Tech Revenue']) * 100,
                               np.nan)
    
    # Merge with Engagement Partner if available (after all calculations)
    if partners is not None:
        grp = grp.merge(partners, on='Engagement ID', how='left')

        # Fill in blank Engagement Partners by finding similar engagement names
        for idx, row in grp.iterrows():
            if pd.isna(row['Engagement Partner']) and row['Engagement Name']:
                eng_name = row['Engagement Name']
                if 'Offshore' in eng_name:
                    similar_name = eng_name.replace('Offshore', 'Onshore')
                elif 'Onshore' in eng_name:
                    similar_name = eng_name.replace('Onshore', 'Offshore')
                else:
                    similar_name = None

                if similar_name:
                    match = grp[grp['Engagement Name'] == similar_name]
                    if not match.empty and pd.notna(match['Engagement Partner'].iloc[0]):
                        grp.at[idx, 'Engagement Partner'] = match['Engagement Partner'].iloc[0]

    # Merge NUI ETD + BoB partner if available
    if nui_etd is not None:
        grp = grp.merge(nui_etd, on='Engagement ID', how='left')

    # Backfill Engagement Partner from BoB if missing
    if 'Engagement Partner' in grp.columns and 'Engagement Partner (BoB)' in grp.columns:
        grp['Engagement Partner'] = grp['Engagement Partner'].fillna(grp['Engagement Partner (BoB)'])
    elif 'Engagement Partner' not in grp.columns and 'Engagement Partner (BoB)' in grp.columns:
        grp['Engagement Partner'] = grp['Engagement Partner (BoB)']

    # Column order
    cols = ['Engagement ID', 'Engagement Name', 'ANSR / Tech Revenue', 'Margin Cost', 'Margin Amount', 'TER', 'Margin %', 'Hours']
    if 'Engagement Partner' in grp.columns:
        cols.append('Engagement Partner')
    if 'Engagement Manager' in grp.columns:
        cols.append('Engagement Manager')
    if 'Engagement Status' in grp.columns:
        cols.append('Engagement Status')
    if 'NUI ETD' in grp.columns:
        cols.append('NUI ETD')

    # Ensure we don't leak the BoB helper column
    grp = grp[[c for c in cols if c in grp.columns]]
    
    grp = grp.round(2).sort_values(by='TER', ascending=False)
    return grp


def add_totals_and_format(ws) -> None:
    """Add totals row and apply comprehensive Excel formatting."""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    headers = [cell.value for cell in ws[1]]
    last_row = ws.max_row
    ws.cell(row=last_row + 1, column=1).value = 'TOTAL'
    eng_numeric_cols = ['ANSR / Tech Revenue', 'Margin Cost', 'Margin Amount', 'TER', 'Hours', 'NUI ETD']
    for col_name in eng_numeric_cols:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            start_cell = f'{col_letter}2'
            end_cell = f'{col_letter}{last_row}'
            ws.cell(row=last_row + 1, column=col_idx).value = f'=SUBTOTAL(9,{start_cell}:{end_cell})'

    if 'NUI ETD' in headers:
        col_idx = headers.index('NUI ETD') + 1
        col_letter = get_column_letter(col_idx)
        data_range = f'{col_letter}2:{col_letter}{last_row}'
        red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
        ws.conditional_formatting.add(data_range, CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill))
        ws.conditional_formatting.add(data_range, CellIsRule(operator='lessThan', formula=['0'], fill=green_fill))


def format_engagement_summary_sheet(ws) -> None:
    """Apply comprehensive formatting to Engagement Summary sheets."""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # Define column widths for Engagement Summary
    column_widths = {
        'Engagement ID': 16.43,
        'Engagement Name': 48.14,
        'ANSR / Tech Revenue': 23.57,
        'Margin Cost': 16.86,
        'Margin Amount': 18.43,
        'TER': 16.86,
        'Margin %': 11.43,
        'Hours': 13.0,
        'Engagement Partner': 21.14,
        'Engagement Manager': 22.43,
        'Engagement Status': 20.0,
        'NUI ETD': 14.29
    }
    
    # Define number formats
    money_format = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Format header row
    thin_border = Border(bottom=Side(style='thin'))
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=11, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = thin_border
        
        # Set column width
        col_letter = get_column_letter(col_idx)
        if header in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[header]
    
    # Format data rows
    last_row = ws.max_row
    for row_idx in range(2, last_row + 1):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Apply number formatting
            if header in ['ANSR / Tech Revenue', 'Margin Cost', 'Margin Amount', 'TER', 'NUI ETD']:
                cell.number_format = money_format


def format_employee_summary_sheet(ws) -> None:
    """Apply comprehensive formatting to Employee Summary sheet."""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # Define column widths for Employee Summary
    column_widths = {
        'Employee / Product Name': 37.29,
        'Employee GUI / Product ID': 26.71,
        'Rank / Method': 17.43,
        'Grade': 8.57,
        'Employee Region': 18.57,
        'Country / Region': 18.0,
        'Service Line': 13.71,
        'Hours': 9.14,
        'NSR': 16.86,
        'ANSR': 13.0,
        'Margin Cost': 13.0,
        'Expense Amount': 19.57,
        '#Engagements': 16.0,
        '#Opportunities': 16.57,
        'TER': 16.86,
        'Margin Amount': 18.43,
        'Margin % (on ANSR)': 20.86,
        'EAF (ANSR/NSR)': 17.57,
        'Level': 7.86
    }
    
    # Define number formats
    money_format = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Format header row
    thin_border = Border(bottom=Side(style='thin'))
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=11, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = thin_border
        
        # Set column width
        col_letter = get_column_letter(col_idx)
        if header in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[header]
    
    # Format data rows
    last_row = ws.max_row
    for row_idx in range(2, last_row + 1):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Apply number formatting
            if header in ['NSR', 'ANSR', 'Margin Cost', 'Expense Amount', 'TER', 'Margin Amount']:
                cell.number_format = money_format


def format_monthly_summary_sheet(ws) -> None:
    """Apply comprehensive formatting to Monthly Summary sheet."""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # Define number formats
    money_format = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Format header row
    thin_border = Border(bottom=Side(style='thin'))
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=11, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = thin_border
        
        # Set column width
        col_letter = get_column_letter(col_idx)
        if header == 'Month':
            ws.column_dimensions[col_letter].width = 12.0
        else:
            ws.column_dimensions[col_letter].width = 16.86
    
    # Format data rows
    last_row = ws.max_row
    for row_idx in range(2, last_row + 1):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Apply number formatting
            if header in ['ANSR / Tech Revenue', 'Margin Cost', 'Margin Amount', 'Expense Amount', 'TER']:
                cell.number_format = money_format


def format_recon_sheet(ws) -> None:
    """Apply comprehensive formatting to WIP vs BoB Recon sheet."""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Format header row
    thin_border = Border(bottom=Side(style='thin'))
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=11, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = thin_border
        
        # Set column width based on content
        col_letter = get_column_letter(col_idx)
        if header == 'Engagement ID':
            ws.column_dimensions[col_letter].width = 16.43
        elif header == 'Engagement Name':
            ws.column_dimensions[col_letter].width = 48.14
        elif header in ['In WIP', 'In BoB']:
            ws.column_dimensions[col_letter].width = 10.0
        elif header == 'Status':
            ws.column_dimensions[col_letter].width = 20.0
        else:
            ws.column_dimensions[col_letter].width = 15.0


def format_kpi_sheet(ws) -> None:
    """Apply comprehensive formatting to KPI sheets."""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # Define number formats
    money_format = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Format header row
    thin_border = Border(bottom=Side(style='thin'))
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=11, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        cell.border = thin_border
        
        # Set column width
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20.0
    
    # Format data row
    if ws.max_row >= 2:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            # Apply money formatting to relevant columns
            if any(term in str(header) for term in ['Amount', 'Revenue', 'Cost', 'BILLINGS', 'ANSR', 'TER', 'Expense', 'Margin']):
                cell.number_format = money_format


def employee_level(rank: str, grade: str) -> float:
    r = (rank or '').strip().lower()
    g = (grade or '').strip()
    if r in ['partner/principal', 'partner']:
        return 7
    if r == 'executive director':
        return 6
    if r == 'senior manager' and g == '2':
        return 5
    if r == 'senior manager' and g == '1':
        return 4
    if r == 'manager':
        return 3
    if r == 'senior':
        return 2
    if r in ['staff', 'assistant', 'staff/assistant']:
        return 1
    return np.nan


def load_engagement_partners(bills_file: Optional[str]) -> Optional[pd.DataFrame]:
    """Load Engagement Partner mapping from Bills file."""
    if bills_file is None:
        return None
    # Try "Engagement Partners" sheet first, then "Billing" sheet (from prepare_bills.py)
    for sheet_name in ['Engagement Partners', 'Billing']:
        try:
            partners = pd.read_excel(bills_file, sheet_name=sheet_name, engine='openpyxl')
            partners.columns = [str(c).strip() for c in partners.columns]
            if 'Engagement ID' in partners.columns:
                # Normalize partner column name
                if 'Billing Partner' in partners.columns and 'Engagement Partner' not in partners.columns:
                    partners.rename(columns={'Billing Partner': 'Engagement Partner'}, inplace=True)
                if 'Engagement Partner' in partners.columns:
                    # If Billing sheet, dedupe to a single partner per Engagement ID
                    cols_to_keep = ['Engagement ID', 'Engagement Partner']
                    if 'Billing Amount' in partners.columns:
                        # Prefer the partner with highest billing amount
                        partners_sorted = partners.sort_values(by='Billing Amount', ascending=False)
                        partners_dedup = partners_sorted.dropna(subset=['Engagement Partner']).drop_duplicates(subset=['Engagement ID'])
                        return partners_dedup[cols_to_keep]
                    # Otherwise, just drop duplicates and keep first non-null partner
                    partners_dedup = partners.dropna(subset=['Engagement Partner']).drop_duplicates(subset=['Engagement ID'])
                    return partners_dedup[cols_to_keep]
        except Exception:
            continue
    print('Warning: Could not load Engagement Partner data from Bills file', file=sys.stderr)
    return None


def load_nui_etd(bob_file: Optional[str]) -> Optional[pd.DataFrame]:
    """Load NUI ETD mapping from BoB file Export sheet."""
    if bob_file is None:
        return None
    try:
        nui = pd.read_excel(bob_file, sheet_name='Export', engine='openpyxl')
        nui.columns = [str(c).strip() for c in nui.columns]
        if 'Engagement ID' in nui.columns:
            cols = ['Engagement ID']
            for c in ['NUI ETD', 'Engagement Manager', 'Engagement Status', 'Engagement Partner', 'Billing Partner']:
                if c in nui.columns:
                    cols.append(c)
            if len(cols) > 1:
                df = nui[cols].copy()
                # Standardize partner column name to a BoB-specific field to avoid merge collisions
                if 'Engagement Partner' in df.columns:
                    df.rename(columns={'Engagement Partner': 'Engagement Partner (BoB)'}, inplace=True)
                elif 'Billing Partner' in df.columns:
                    df.rename(columns={'Billing Partner': 'Engagement Partner (BoB)'}, inplace=True)
                return df
        return None
    except Exception as e:
        print(f'Warning: Could not load NUI ETD data: {e}', file=sys.stderr)
        return None


def wip_bob_reconciliation(df_wip: pd.DataFrame, bob_data: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Create reconciliation of Engagement IDs between WIP and BoB."""
    # Get unique engagement IDs and names from WIP
    wip_engagements = df_wip[['Engagement ID', 'Engagement Name']].drop_duplicates()
    wip_ids = set(wip_engagements['Engagement ID'].dropna().unique())
    
    # Get unique engagement IDs from BoB
    bob_ids = set()
    if bob_data is not None and 'Engagement ID' in bob_data.columns:
        bob_ids = set(bob_data['Engagement ID'].dropna().unique())
    
    # Find differences
    only_in_wip = wip_ids - bob_ids
    only_in_bob = bob_ids - wip_ids
    in_both = wip_ids & bob_ids
    
    # Build reconciliation dataframe
    recon_data = []
    
    # Engagements in both
    for eng_id in sorted(in_both):
        eng_name = wip_engagements[wip_engagements['Engagement ID'] == eng_id]['Engagement Name'].iloc[0] if not wip_engagements[wip_engagements['Engagement ID'] == eng_id].empty else ''
        recon_data.append({
            'Engagement ID': eng_id,
            'Engagement Name': eng_name,
            'In WIP': 'Yes',
            'In BoB': 'Yes',
            'Status': 'Matched'
        })
    
    # Engagements only in WIP
    for eng_id in sorted(only_in_wip):
        eng_name = wip_engagements[wip_engagements['Engagement ID'] == eng_id]['Engagement Name'].iloc[0] if not wip_engagements[wip_engagements['Engagement ID'] == eng_id].empty else ''
        recon_data.append({
            'Engagement ID': eng_id,
            'Engagement Name': eng_name,
            'In WIP': 'Yes',
            'In BoB': 'No',
            'Status': 'Missing in BoB'
        })
    
    # Engagements only in BoB
    for eng_id in sorted(only_in_bob):
        recon_data.append({
            'Engagement ID': eng_id,
            'Engagement Name': '',
            'In WIP': 'No',
            'In BoB': 'Yes',
            'Status': 'Missing in WIP'
        })
    
    recon_df = pd.DataFrame(recon_data)
    
    # Add summary at the top
    summary_data = [
        {'Engagement ID': 'SUMMARY', 'Engagement Name': '', 'In WIP': '', 'In BoB': '', 'Status': ''},
        {'Engagement ID': f'Total in WIP: {len(wip_ids)}', 'Engagement Name': '', 'In WIP': '', 'In BoB': '', 'Status': ''},
        {'Engagement ID': f'Total in BoB: {len(bob_ids)}', 'Engagement Name': '', 'In WIP': '', 'In BoB': '', 'Status': ''},
        {'Engagement ID': f'Matched: {len(in_both)}', 'Engagement Name': '', 'In WIP': '', 'In BoB': '', 'Status': ''},
        {'Engagement ID': f'Only in WIP: {len(only_in_wip)}', 'Engagement Name': '', 'In WIP': '', 'In BoB': '', 'Status': ''},
        {'Engagement ID': f'Only in BoB: {len(only_in_bob)}', 'Engagement Name': '', 'In WIP': '', 'In BoB': '', 'Status': ''},
        {'Engagement ID': '', 'Engagement Name': '', 'In WIP': '', 'In BoB': '', 'Status': ''},
    ]
    
    summary_df = pd.DataFrame(summary_data)
    result = pd.concat([summary_df, recon_df], ignore_index=True)
    
    return result


def employee_summary(df: pd.DataFrame) -> pd.DataFrame:
    grp = df.groupby(['Employee / Product Name', 'Employee GUI / Product ID', 'Rank / Method', 'Grade',
                      'Employee Region', 'Country / Region', 'Service Line']).agg({
        'Charged Hours / Quantity': 'sum',
        'NSR / Tech Revenue': 'sum',
        'ANSR / Tech Revenue': 'sum',
        'Margin Cost': 'sum',
        'Expense Amount': 'sum',
        'Opportunity ID': pd.Series.nunique,
        'Engagement ID': pd.Series.nunique
    }).reset_index()

    grp.rename(columns={
        'Charged Hours / Quantity': 'Hours',
        'NSR / Tech Revenue': 'NSR',
        'ANSR / Tech Revenue': 'ANSR',
        'Opportunity ID': '#Opportunities',
        'Engagement ID': '#Engagements'
    }, inplace=True)

    grp['TER'] = grp['ANSR'] + grp['Expense Amount']
    grp['Margin Amount'] = grp['ANSR'] - grp['Margin Cost']
    grp['Margin % (on ANSR)'] = np.where(grp['ANSR'] != 0, (grp['Margin Amount'] / grp['ANSR']) * 100, np.nan)
    grp['EAF (ANSR/NSR)'] = np.where(grp['NSR'] != 0, grp['ANSR'] / grp['NSR'], np.nan)
    grp['Level'] = [employee_level(r, g) for r, g in zip(grp['Rank / Method'], grp['Grade'])]

    grp = grp[['Employee / Product Name', 'Employee GUI / Product ID', 'Rank / Method', 'Grade', 'Employee Region',
               'Country / Region', 'Service Line', 'Hours', 'NSR', 'ANSR', 'Margin Cost', 'Expense Amount',
               '#Engagements', '#Opportunities', 'TER', 'Margin Amount', 'Margin % (on ANSR)', 'EAF (ANSR/NSR)', 'Level']]
    grp = grp.round(2).sort_values(by=['Level', 'TER'], ascending=[False, False])
    return grp


def kpi_totals(df: pd.DataFrame) -> dict:
    ansr_total = float(df['ANSR / Tech Revenue'].sum())
    margin_cost_total = float(df['Margin Cost'].sum())
    expense_total = float(df['Expense Amount'].sum())
    ter_total = ansr_total + expense_total
    margin_amount_total = ansr_total - margin_cost_total
    margin_pct_overall = (margin_amount_total / ansr_total * 100.0) if ansr_total != 0 else float('nan')
    hours_total = float(df['Charged Hours / Quantity'].sum())
    return {
        'Hours_total': round(hours_total, 2),
        'ANSR_total': round(ansr_total, 2),
        'Expense_total': round(expense_total, 2),
        'TER_total': round(ter_total, 2),
        'MarginCost_total': round(margin_cost_total, 2),
        'MarginAmount_total': round(margin_amount_total, 2),
        'MarginPct_overall_on_ANSR': round(margin_pct_overall, 2)
    }


def kpi_bridge(ansr_total: float, ter_total: float, margin_amount_total: float,
               billings: float, target_margin_pct: float) -> dict:
    write_on = billings - ter_total
    t = target_margin_pct / 100.0
    additional_rev_needed = (t * ansr_total - margin_amount_total) / (1.0 - t)
    revised_margin_amount = margin_amount_total + additional_rev_needed
    revised_revenue = ansr_total + additional_rev_needed
    revised_margin_pct = (revised_margin_amount / revised_revenue * 100.0) if revised_revenue != 0 else float('nan')
    gap_pct = (additional_rev_needed / revised_revenue * 100.0) if revised_revenue != 0 else float('nan')

    result = {
        'BILLINGS': round(billings, 2),
        'Write_on': round(write_on, 2),
        'Revised_Margin_Amount': round(revised_margin_amount, 2),
        'Revised_Revenue': round(revised_revenue, 2),
        'Revised_Margin_%': round(revised_margin_pct, 2),
        'Additional_Revenue_Needed': round(additional_rev_needed, 2),
        'Gap_%_of_Revised_Revenue': round(gap_pct, 2),
    }
    if write_on > 0:
        result['WriteOn_Remaining'] = round(write_on - additional_rev_needed, 2)
    else:
        result['Write_Off_Plus_Additional_Revenue_Required'] = round(abs(write_on) + additional_rev_needed, 2)
    return result


def build_markdown_table(df: pd.DataFrame) -> str:
    try:
        return df.to_markdown(index=False)
    except Exception:
        cols = df.columns.tolist()
        header = '|' + '|'.join(cols) + '|\n'
        sep = '|' + '|'.join(['---'] * len(cols)) + '|\n'
        rows = []
        for _, r in df.head(10).iterrows():
            rows.append('|' + '|'.join(str(r[c]) for c in cols) + '|\n')
        return header + sep + ''.join(rows)


def run(input_file: str,
    detail_sheet: str,
    header_row_index: Optional[int],
    fy_start: str,
    fy_end: str,
    billings_str: Optional[str],
    target_margin_pct: Optional[float],
    output_file: str,
    bills_file: Optional[str] = None,
    bob_file: Optional[str] = None,
    print_markdown: bool = False) -> Tuple[pd.DataFrame, pd.DataFrame, dict, Optional[dict]]:

    df_raw = load_detail_frame(input_file, detail_sheet, header_row_index)

    numeric_cols = ['Charged Hours / Quantity', 'NSR / Tech Revenue', 'ANSR / Tech Revenue', 'Margin Cost', 'Expense Amount']
    for c in numeric_cols:
        if c not in df_raw.columns:
            df_raw[c] = 0

    df_all = coerce_numeric(df_raw.copy(), numeric_cols)
    df_fy = filter_fiscal_year(df_all.copy(), pd.Timestamp(fy_start), pd.Timestamp(fy_end))

    # Monthly summary (FY window)
    date_col = 'Accounting Date' if 'Accounting Date' in df_fy.columns and df_fy['Accounting Date'].notna().any() else 'Transaction Date'
    df_fy['Month'] = df_fy[date_col].dt.to_period('M').astype(str)
    monthly = df_fy.groupby('Month').agg({
        'Charged Hours / Quantity': 'sum',
        'ANSR / Tech Revenue': 'sum',
        'Margin Cost': 'sum',
        'Expense Amount': 'sum'
    }).reset_index()
    monthly.rename(columns={'Charged Hours / Quantity': 'Hours'}, inplace=True)
    monthly['Margin Amount'] = monthly['ANSR / Tech Revenue'] - monthly['Margin Cost']
    monthly['TER'] = monthly['ANSR / Tech Revenue'] + monthly['Expense Amount']
    monthly['Margin %'] = np.where(monthly['ANSR / Tech Revenue'] != 0, (monthly['Margin Amount'] / monthly['ANSR / Tech Revenue']) * 100, np.nan)
    monthly = monthly.round(2).sort_values('Month')

    # Load engagement partners and NUI ETD
    partners = load_engagement_partners(bills_file)
    nui_etd = load_nui_etd(bob_file)

    eng_fy = engagement_summary(df_fy, partners, nui_etd)
    eng_etd = engagement_summary(df_all, partners, nui_etd)
    emp = employee_summary(df_fy)
    totals = kpi_totals(df_fy)

    # Create WIP vs BoB reconciliation
    recon = wip_bob_reconciliation(df_all, nui_etd)

    bridge = None
    if billings_str is not None and target_margin_pct is not None:
        billings = parse_billings(billings_str)
        bridge = kpi_bridge(ansr_total=float(totals['ANSR_total']),
                            ter_total=float(totals['TER_total']),
                            margin_amount_total=float(totals['MarginAmount_total']),
                            billings=billings,
                            target_margin_pct=float(target_margin_pct))

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        eng_fy.to_excel(writer, sheet_name='Engagement Summary (FYTD)', index=False)
        eng_etd.to_excel(writer, sheet_name='Engagement Summary (ETD)', index=False)
        emp.to_excel(writer, sheet_name='Employee Summary', index=False)
        monthly.to_excel(writer, sheet_name='Monthly Summary', index=False)
        recon.to_excel(writer, sheet_name='WIP vs BoB Recon', index=False)
        pd.DataFrame([totals]).to_excel(writer, sheet_name='KPI Totals', index=False)
        if bridge is not None:
            pd.DataFrame([bridge]).to_excel(writer, sheet_name='KPI Bridge', index=False)

    # Add total rows with SUBTOTAL formulas and apply comprehensive formatting
    wb = load_workbook(output_file)

    # Engagement Summary totals + conditional formatting (both FYTD and ETD)
    for sheet_name in ['Engagement Summary (FYTD)', 'Engagement Summary (ETD)']:
        if sheet_name in wb.sheetnames:
            add_totals_and_format(wb[sheet_name])
            format_engagement_summary_sheet(wb[sheet_name])

    # Employee Summary totals and formatting
    if 'Employee Summary' in wb.sheetnames:
        ws_emp = wb['Employee Summary']
        last_row_emp = ws_emp.max_row
        headers_emp = [cell.value for cell in ws_emp[1]]
        ws_emp.cell(row=last_row_emp + 1, column=1).value = 'TOTAL'
        emp_numeric_cols = ['Hours', 'NSR', 'ANSR', 'Margin Cost', 'Expense Amount', '#Engagements', '#Opportunities', 'TER', 'Margin Amount']
        for col_name in emp_numeric_cols:
            if col_name in headers_emp:
                col_idx = headers_emp.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                start_cell = f'{col_letter}2'
                end_cell = f'{col_letter}{last_row_emp}'
                formula = f'=SUBTOTAL(9,{start_cell}:{end_cell})'
                ws_emp.cell(row=last_row_emp + 1, column=col_idx).value = formula
        format_employee_summary_sheet(ws_emp)

    # Monthly Summary totals and formatting
    if 'Monthly Summary' in wb.sheetnames:
        ws_monthly = wb['Monthly Summary']
        last_row_monthly = ws_monthly.max_row
        headers_monthly = [cell.value for cell in ws_monthly[1]]
        ws_monthly.cell(row=last_row_monthly + 1, column=1).value = 'TOTAL'
        monthly_numeric_cols = ['Hours', 'ANSR / Tech Revenue', 'Margin Cost', 'Expense Amount', 'Margin Amount', 'TER']
        for col_name in monthly_numeric_cols:
            if col_name in headers_monthly:
                col_idx = headers_monthly.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                start_cell = f'{col_letter}2'
                end_cell = f'{col_letter}{last_row_monthly}'
                formula = f'=SUBTOTAL(9,{start_cell}:{end_cell})'
                ws_monthly.cell(row=last_row_monthly + 1, column=col_idx).value = formula
        format_monthly_summary_sheet(ws_monthly)
    
    if 'WIP vs BoB Recon' in wb.sheetnames:
        format_recon_sheet(wb['WIP vs BoB Recon'])
    
    if 'KPI Totals' in wb.sheetnames:
        format_kpi_sheet(wb['KPI Totals'])
    
    if 'KPI Bridge' in wb.sheetnames:
        format_kpi_sheet(wb['KPI Bridge'])

    wb.save(output_file)

    print('\n=== KPI TOTALS ===')
    print(pd.DataFrame([totals]).to_string(index=False))

    if bridge is not None:
        print('\n=== KPI BRIDGE (Billings & Target Margin) ===')
        print(pd.DataFrame([bridge]).to_string(index=False))

    if print_markdown:
        print('\n=== Engagement Summary (Top 25 by TER) [Markdown] ===')
        md = build_markdown_table(eng_fy.head(25))
        print(md)

    return eng_fy, emp, totals, bridge


def main():
    parser = argparse.ArgumentParser(description='FY Engagement & Employee Summary + KPI Bridge')
    parser.add_argument('--input', required=True, help='Input WIP Excel file (e.g., WIPs_20260115.xlsx)')
    parser.add_argument('--sheet', default='Detail', help='Transactional sheet name')
    parser.add_argument('--header-row-index', type=int, default=None, help='0-based header row index (if omitted, auto-detect)')
    parser.add_argument('--fy-start', required=True, help='Fiscal start date (YYYY-MM-DD)')
    parser.add_argument('--fy-end', required=True, help='Fiscal end date (YYYY-MM-DD)')
    parser.add_argument('--billings', default=None, help='Billings total, supports 20.11M or plain numbers')
    parser.add_argument('--target-margin-pct', type=float, default=None, help='Target margin percent (e.g., 29)')
    parser.add_argument('--output', required=True, help='Output Excel file')
    parser.add_argument('--bills', default=None, help='Bills file with Engagement Partner data (e.g., Bills_08.12.2025.xlsx)')
    parser.add_argument('--bob', default=None, help='BoB file with NUI ETD data on sheet Export (e.g., BoB_08.12.2025.xlsx)')
    parser.add_argument('--print-markdown', action='store_true', help='Print markdown table for top engagements')

    args = parser.parse_args()

    if args.billings is None:
        args.billings = input('Enter Billings total (supports 20.11M or plain numbers): ').strip()

    try:
        run(input_file=args.input,
            detail_sheet=args.sheet,
            header_row_index=args.header_row_index,
            fy_start=args.fy_start,
            fy_end=args.fy_end,
            billings_str=args.billings,
            target_margin_pct=args.target_margin_pct,
            output_file=args.output,
            bills_file=args.bills,
            bob_file=args.bob,
            print_markdown=args.print_markdown)
    except Exception as e:
        print(f'Error: {e}', file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
   
