# -*- coding: utf-8 -*-
"""
Prepare BoB*.xlsx Input File
-----------------------------
Automates data preparation for BoB (Book of Business) file before running engagement analysis.

Steps:
1. Extract "Engagement ID" from "Engagement Name (ID) Currency" column in Export sheet
2. Preserve existing columns (NUI ETD, Engagement Manager, Engagement Status)

Usage:
    python prepare_bob.py --input BoB_08.12.2025.xlsx --output BoB_08.12.2025_prepared.xlsx
"""

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def extract_engagement_id(name_str: str) -> str:
    """Extract Engagement ID from 'Engagement Name (ID) Currency' format.
    
    Example: "Project XYZ (E-87654321) USD" -> "E-87654321"
    """
    if pd.isna(name_str):
        return None
    
    # Look for pattern (E-XXXXXXXX) or similar
    match = re.search(r'\(([A-Z]-\d+)\)', str(name_str))
    if match:
        return match.group(1)
    return None


def prepare_bob_export(input_file: str, output_file: str) -> None:
    """Prepare BoB Export sheet with Engagement ID extraction."""
    
    print(f"Loading {input_file}...")
    
    # Read Export sheet
    try:
        df = pd.read_excel(input_file, sheet_name='Export', engine='openpyxl')
    except Exception as e:
        print(f"Error reading Export sheet: {e}", file=sys.stderr)
        sys.exit(1)
    
    # Strip column names
    df.columns = [str(c).strip() for c in df.columns]
    
    print(f"Found {len(df)} rows in Export sheet")
    
    # Extract Engagement ID
    if 'Engagement Name (ID) Currency' not in df.columns:
        print("Warning: 'Engagement Name (ID) Currency' column not found.", file=sys.stderr)
        print("Available columns:", df.columns.tolist())
        
        # Try alternative column names
        possible_cols = [c for c in df.columns if 'engagement' in c.lower() and 'name' in c.lower()]
        if possible_cols:
            print(f"Found possible alternative: {possible_cols[0]}")
            source_col = possible_cols[0]
        else:
            print("Error: Cannot find engagement name column for ID extraction", file=sys.stderr)
            sys.exit(1)
    else:
        source_col = 'Engagement Name (ID) Currency'
    
    # Check if Engagement ID already exists
    if 'Engagement ID' in df.columns:
        print("ℹ 'Engagement ID' column already exists - will overwrite with extracted values")
    
    df['Engagement ID'] = df[source_col].apply(extract_engagement_id)
    extracted_count = df['Engagement ID'].notna().sum()
    print(f"✓ Extracted Engagement ID for {extracted_count}/{len(df)} rows")
    
    # Verify required columns for analysis
    required_cols = ['Engagement ID', 'NUI ETD', 'Engagement Manager', 'Engagement Status']
    found_cols = [c for c in required_cols if c in df.columns]
    missing_cols = [c for c in required_cols if c not in df.columns]
    
    print(f"\nColumns for analysis:")
    for col in found_cols:
        non_null = df[col].notna().sum()
        print(f"  ✓ {col}: {non_null}/{len(df)} populated")
    
    if missing_cols:
        print(f"\nWarning: Missing columns: {missing_cols}")
        print("  These will be unavailable in the engagement analysis")
    
    # Write to output file
    print(f"\nWriting to {output_file}...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write updated Export sheet
        df.to_excel(writer, sheet_name='Export', index=False)
        
        # Copy other sheets from original file
        try:
            original_wb = load_workbook(input_file, read_only=False)
            for sheet_name in original_wb.sheetnames:
                if sheet_name != 'Export':
                    original_sheet = original_wb[sheet_name]
                    new_sheet = writer.book.create_sheet(sheet_name)
                    
                    for row in original_sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet[cell.coordinate]
                            new_cell.value = cell.value
                            if cell.has_style:
                                new_cell.font = cell.font.copy()
                                new_cell.border = cell.border.copy()
                                new_cell.fill = cell.fill.copy()
                                new_cell.number_format = cell.number_format
                                new_cell.protection = cell.protection.copy()
                                new_cell.alignment = cell.alignment.copy()
        except Exception as e:
            print(f"Warning: Could not copy other sheets: {e}")
    
    print(f"✓ Successfully prepared {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Prepare BoB*.xlsx input file for engagement analysis')
    parser.add_argument('--input', required=True, help='Input BoB Excel file (e.g., BoB_08.12.2025.xlsx)')
    parser.add_argument('--output', required=True, help='Output prepared Excel file (e.g., BoB_08.12.2025_prepared.xlsx)')
    
    args = parser.parse_args()
    
    if not Path(args.input).exists():
        print(f"Error: Input file '{args.input}' not found", file=sys.stderr)
        sys.exit(1)
    
    prepare_bob_export(args.input, args.output)
    
    print("\n" + "="*60)
    print("NEXT STEPS:")
    print("="*60)
    print(f"Use this file in your analysis: --bob {args.output}")


if __name__ == '__main__':
    main()
